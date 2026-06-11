import json
import hashlib
import io
import random
import re
from typing import Any, Optional

from openpyxl import load_workbook

from app.utils.cache_manager import cache_manager
from .quotations_schema import QuotationItem


class QuotationsService:
	CATALOG_KEY = "quotations:catalog"

	def _ensure_redis(self):
		if not cache_manager.redis_client:
			raise RuntimeError("Redis is not available")
		return cache_manager.redis_client

	def _history_key(self, user_id: str) -> str:
		return f"quotations_history:{user_id}"

	def _normalize_text(self, value: Any) -> str:
		if value is None:
			return ""
		return str(value).strip()

	def _normalize_column_name(self, value: Any) -> str:
		text = self._normalize_text(value).lower()
		text = re.sub(r"[^a-z0-9]+", "_", text)
		return text.strip("_")

	def _normalize_row(self, row: dict[str, Any]) -> dict[str, Any]:
		normalized_row: dict[str, Any] = {}
		for key, value in row.items():
			normalized_row[self._normalize_column_name(key)] = value
		return normalized_row

	def _cell_value(self, value: Any) -> Any:
		if value is None:
			return ""
		return value

	def _row_has_meaningful_data(self, row: list[Any]) -> bool:
		return any(self._normalize_text(cell) for cell in row)

	def _detect_header_row(self, rows: list[list[Any]]) -> tuple[int, list[str]]:
		required_markers = {"goal", "quote"}
		for index, row in enumerate(rows[:25]):
			normalized_headers = [self._normalize_column_name(cell) for cell in row]
			non_empty_headers = [header for header in normalized_headers if header]
			if not non_empty_headers:
				continue

			marker_set = set(non_empty_headers)
			if required_markers.issubset(marker_set):
				return index, normalized_headers

			if "quote" in marker_set and len(marker_set.intersection({"goal", "attribution_display", "religious_filter", "content_type"})) >= 1:
				return index, normalized_headers

		raise RuntimeError("Could not detect a header row in the uploaded worksheet")

	def _build_records_from_sheet(self, excel_bytes: bytes) -> tuple[list[dict[str, Any]], list[str], str]:
		workbook = load_workbook(io.BytesIO(excel_bytes), data_only=True, read_only=True)
		worksheet = workbook[workbook.sheetnames[0]]
		rows = worksheet.iter_rows(values_only=True)
		buffered_rows: list[list[Any]] = []
		headers_row_index = -1
		headers: list[str] = []

		for index, row in enumerate(rows):
			row_values = [self._cell_value(cell) for cell in row]
			buffered_rows.append(row_values)
			if index >= 24:
				headers_row_index, headers = self._detect_header_row(buffered_rows)
				break

		if headers_row_index == -1:
			headers_row_index, headers = self._detect_header_row(buffered_rows)

		records: list[dict[str, Any]] = []

		for row in buffered_rows[headers_row_index + 1 :]:
			if not self._row_has_meaningful_data(row):
				continue

			record: dict[str, Any] = {}
			for index, header in enumerate(headers):
				if not header:
					continue
				value = row[index] if index < len(row) else None
				record[header] = value
			records.append(record)

		for row in rows:
			row_values = [self._cell_value(cell) for cell in row]
			if not self._row_has_meaningful_data(row_values):
				continue

			record = {}
			for index, header in enumerate(headers):
				if not header:
					continue
				value = row_values[index] if index < len(row_values) else None
				record[header] = value
			records.append(record)

		return records, headers, worksheet.title

	def _parse_tags(self, value: Any) -> list[str] | None:
		text = self._normalize_text(value)
		if not text:
			return None
		return [item.strip() for item in text.split(",") if item.strip()]

	def _row_to_item(self, row: dict[str, Any]) -> QuotationItem:
		normalized_row = self._normalize_row(row)
		quote_text = self._normalize_text(normalized_row.get("quote"))
		attribution = self._normalize_text(normalized_row.get("attribution_display"))
		source_work = self._normalize_text(normalized_row.get("source_work_or_reference"))
		goal = self._normalize_text(normalized_row.get("goal"))
		source_genre = self._normalize_text(normalized_row.get("source_genre"))
		content_type = self._normalize_text(normalized_row.get("content_type"))
		quote_id_source = "|".join(
			[goal, quote_text, attribution, source_work, source_genre, self._normalize_text(normalized_row.get("religious_filter"))]
		)
		quote_id = hashlib.sha1(quote_id_source.encode("utf-8")).hexdigest()

		return QuotationItem(
			id=quote_id,
			goal=goal or None,
			quote=quote_text,
			attribution_display=attribution or None,
			content_type=content_type or None,
			source_genre=source_genre or None,
			source_work_or_reference=source_work or None,
			risk_number=self._to_int(normalized_row.get("risk")),
			filter=self._normalize_text(normalized_row.get("filter")) or None,
			goal_tags=self._parse_tags(normalized_row.get("goal_tags")),
			intensity=self._to_int(normalized_row.get("intensity")),
			religious_filter=self._normalize_text(normalized_row.get("religious_filter")) or None,
			action_style=self._normalize_text(normalized_row.get("action_style")) or None,
			energy_type=self._normalize_text(normalized_row.get("energy_type")) or None,
			notes=self._normalize_text(normalized_row.get("notes")) or None,
		)

	def _to_int(self, value: Any) -> Optional[int]:
		if value is None or value == "":
			return None
		try:
			return int(float(value))
		except Exception:
			return None

	def upload_excel(self, excel_bytes: bytes) -> int:
		redis_client = self._ensure_redis()
		records, headers, sheet_name = self._build_records_from_sheet(excel_bytes)
		print(f"[QUOTATIONS] Uploaded sheet: {sheet_name}")
		print(f"[QUOTATIONS] Detected headers: {headers}")
		print(f"[QUOTATIONS] Total raw rows: {len(records)}")
		if records:
			first_row = records[0]
			print(f"[QUOTATIONS] First normalized row keys: {list(first_row.keys())}")
			print(f"[QUOTATIONS] First normalized row sample: {first_row}")

		items = []
		filtered_out = 0
		for record in records:
			normalized_record = self._normalize_row(record)
			if not self._normalize_text(normalized_record.get("quote")):
				filtered_out += 1
				continue
			items.append(self._row_to_item(record).model_dump())

		print(f"[QUOTATIONS] Filtered out rows without quote text: {filtered_out}")
		print(f"[QUOTATIONS] Parsed quote items: {len(items)}")
		if items:
			print(f"[QUOTATIONS] First parsed item: {items[0]}")
		else:
			print("[QUOTATIONS] No items parsed from the uploaded file")
		redis_client.set(self.CATALOG_KEY, json.dumps(items))
		return len(items)

	def get_all_quotes(self) -> list[QuotationItem]:
		redis_client = self._ensure_redis()
		raw = redis_client.get(self.CATALOG_KEY)
		if not raw:
			return []
		payload = raw.decode("utf-8") if isinstance(raw, bytes) else str(raw)
		data = json.loads(payload)
		return [QuotationItem(**item) for item in data]

	def get_total_quotes(self) -> int:
		return len(self.get_all_quotes())

	def _normalize_religious_preference(self, value: Optional[str]) -> str:
		text = self._normalize_text(value).lower()
		if not text:
			return ""
		if "christ" in text:
			return "Christian"
		if "spirit" in text:
			return "Spiritual"
		if text in {"secular", "non-religious", "nonreligious", "agnostic"}:
			return "Secular"
		return value.strip()

	def _normalize_goal(self, value: Optional[str]) -> str:
		text = self._normalize_text(value)
		return text.lower()

	def _matches_goal(self, quote: QuotationItem, goal: str) -> bool:
		if not goal:
			return False

		goal_lower = goal.lower()
		candidate_values = [
			self._normalize_text(quote.goal),
			self._normalize_text(quote.filter),
			self._normalize_text(quote.source_genre),
			self._normalize_text(quote.action_style),
			self._normalize_text(quote.energy_type),
			" ".join(quote.goal_tags or []),
		]
		joined = " | ".join(candidate_values).lower()
		return goal_lower in joined

	def _matches_religion(self, quote: QuotationItem, religious_preference: Optional[str]) -> bool:
		preference = self._normalize_religious_preference(religious_preference)
		if not preference:
			return False
		return self._normalize_text(quote.religious_filter).lower() == preference.lower()

	def _prioritized_pools(
		self,
		quotes: list[QuotationItem],
		goal: Optional[str],
		religious_preference: Optional[str],
		seen_ids: set[str],
	) -> list[list[QuotationItem]]:
		goal_value = self._normalize_goal(goal)
		pref_value = self._normalize_religious_preference(religious_preference)
		unseen_quotes = [quote for quote in quotes if quote.id not in seen_ids]

		both = [quote for quote in unseen_quotes if self._matches_goal(quote, goal_value) and self._matches_religion(quote, pref_value)]
		goal_only = [quote for quote in unseen_quotes if self._matches_goal(quote, goal_value) and not self._matches_religion(quote, pref_value)]
		religion_only = [quote for quote in unseen_quotes if not self._matches_goal(quote, goal_value) and self._matches_religion(quote, pref_value)]
		others = [quote for quote in unseen_quotes if not self._matches_goal(quote, goal_value) and not self._matches_religion(quote, pref_value)]

		return [both, goal_only, religion_only, others]

	def _get_user_history(self, user_id: str) -> list[str]:
		redis_client = self._ensure_redis()
		raw = redis_client.lrange(self._history_key(user_id), 0, -1)
		history = []
		for item in raw:
			value = item.decode("utf-8") if isinstance(item, bytes) else str(item)
			history.append(value)
		return history

	def _set_user_history(self, user_id: str, quote_ids: list[str]) -> None:
		redis_client = self._ensure_redis()
		key = self._history_key(user_id)
		redis_client.delete(key)
		if quote_ids:
			redis_client.rpush(key, *quote_ids)

	def _select_candidate(
		self,
		quotes: list[QuotationItem],
		seen_ids: set[str],
		goal: Optional[str],
		religious_preference: Optional[str],
	) -> tuple[QuotationItem, bool]:
		for pool in self._prioritized_pools(quotes, goal, religious_preference, seen_ids):
			if pool:
				return random.choice(pool), False

		return random.choice(quotes), True

	def get_next_quote(self, user_id: str, goal: Optional[str] = None, religious_preference: Optional[str] = None) -> tuple[QuotationItem, int, bool]:
		quotes = self.get_all_quotes()
		if not quotes:
			raise RuntimeError("No quotations have been uploaded yet")

		history = self._get_user_history(user_id)
		seen_ids = set(history)
		selected_quote, exhausted_cycle = self._select_candidate(quotes, seen_ids, goal, religious_preference)

		if exhausted_cycle:
			history = []
			seen_ids = set()

		if selected_quote.id not in seen_ids:
			history.append(selected_quote.id)
		else:
			history = [selected_quote.id]

		self._set_user_history(user_id, history)
		return selected_quote, len(history), exhausted_cycle
